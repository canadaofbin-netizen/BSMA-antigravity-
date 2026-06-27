import sys
import os
import json
import openpyxl

sys.path.append(os.path.abspath(".agents/scripts"))
from universal_excel_inserter import insert_data

data = {
    "article_id": "BSMA0002",
    "inclusion_status": 0,
    "title": "It is ok to be interrupted; it is my job",
    "publication_name": "Qualitative Research in Organizations and Management: An International Journal",
    "author": "De Alwis et al.",
    "year": 2022,
    "exclusion_reason": "Qualitative study; no quantitative boundary spanning data"
}

excel_file = "BSMA_Master_Coding_Sheet.xlsx"
insert_data(excel_file, data)

# Verify
wb = openpyxl.load_workbook(excel_file)
ws = wb.active
for r in range(4, ws.max_row + 1):
    if str(ws.cell(row=r, column=2).value).strip() == "BSMA0002":
        print(f"Verification - Row {r}:")
        print(f"Title: {ws.cell(row=r, column=8).value}")
        print(f"Inclusion/Exclusion Status: {ws.cell(row=r, column=5).value}")
        print(f"Exclusion Reason: {ws.cell(row=r, column=6).value}")
