import openpyxl

wb = openpyxl.load_workbook('BSMA_Master_Coding_Sheet.xlsx')
ws = wb.active

print("Verifying inserted rows for BSMA0008:")
for r in range(4, ws.max_row + 1):
    if ws.cell(row=r, column=2).value == 'BSMA0008':
        row_data = [ws.cell(row=r, column=c).value for c in range(1, 50)]
        print(f"Row {r}: Sample {row_data[2]}, BS: {row_data[39]}, Non-BS: {row_data[43]}, r: {row_data[48]}")
